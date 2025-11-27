from django.shortcuts import render, redirect, get_object_or_404
from dispositivos.models import Usuario, Producto, Proveedor, ProductoProveedor
from dispositivos.forms import UsuarioForm, ProveedorForm, ProductoForm, ProductoProveedorForm, PerfilUsuarioForm, PasswordChangeForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth import update_session_auth_hash
from django.core.paginator import Paginator
from django.utils import timezone
from django.db.models import Q
from django.contrib import messages
from django.http import HttpResponse
import openpyxl

# ------------
# Categorias para editar
# -------------

CATEGORIAS = [
    ("Chocolates", "Chocolates"),
    ("Caramelos", "Caramelos"),
    ("Galletas", "Galletas"),
    ("Gomitas", "Gomitas"),
    ("Alfajores", "Alfajores"),
    ("Otras", "Otra...")
]

from accounts.permissions import (
    require_module_permission,
    get_user_role_name,
    can_edit_own_profile,
    can_assign_roles,
    user_can_add_module,
    user_can_change_module,
    user_can_delete_module
)

# -------------------------------------------------------------
# DASHBOARD
# -------------------------------------------------------------
@login_required
def dashboard(request):
    visitas = request.session.get('visitas', 0)
    request.session['visitas'] = visitas + 1

    context = {
        'visitas': visitas,
        'usuarios_count': Usuario.objects.count(),
        'productos_count': Producto.objects.count(),
        'proveedores_count': Proveedor.objects.count(),
        'movimientos_count': ProductoProveedor.objects.count(),
        'ultimos_usuarios': Usuario.objects.order_by('-idUsuario')[:3],
        'role_name': get_user_role_name(request.user),
    }
    return render(request, "dispositivos/dashboard.html", context)


# -------------------------------------------------------------
# USUARIOS
# -------------------------------------------------------------
@login_required
@require_module_permission('usuarios', 'view')
def formularioUsuario(request):
    visitas = request.session.get("visitas", 0)
    request.session['visitas'] = visitas + 1

    role_name = get_user_role_name(request.user)
    is_admin = role_name == 'administrador'
    is_analyst = role_name == 'analista_financiero'

    # Búsqueda y filtros
    search = request.GET.get("buscar", "")
    rol_filter = request.GET.get("rol", "")
    estado_filter = request.GET.get("estado", "")

    usuarios = Usuario.objects.all()
    if search:
        usuarios = usuarios.filter(Q(username__icontains=search) | Q(nombre__icontains=search))
    if rol_filter:
        usuarios = usuarios.filter(rol__iexact=rol_filter)
    if estado_filter:
        usuarios = usuarios.filter(estado__iexact=estado_filter)

    # ORDENAMIENTO - mECH
    VALID_FIELDS = ['username', 'email', 'nombre', 'apellido', 'rol', 'estado', 'mfa_habilitado']
    order_by = request.GET.get('order_by')
    order = request.GET.get('order')
    if order_by not in VALID_FIELDS or not order_by:
        order_by = 'username'
    if order not in ['asc', 'desc']:
        order = 'asc'
    usuarios = usuarios.order_by(f'-{order_by}' if order == 'desc' else order_by)
    if 'order_by' in request.GET:
        request.session['usuarios_order_by'] = order_by
    if 'order' in request.GET:
        request.session['usuarios_order'] = order

    # Paginado
    pag_size = request.GET.get('pag_size') or request.session.get('usuarios_pag_size', '15')
    if pag_size not in ['5', '15', '30', '100']:
        pag_size = '15'
    request.session['usuarios_pag_size'] = pag_size
    paginator = Paginator(usuarios, int(pag_size))
    page_number = request.GET.get('page')
    usuarios_page = paginator.get_page(page_number)

    # Exportar a Excel
    if "export_excel" in request.GET:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Usuarios"
        headers = ["Username", "Email", "Nombre", "Apellido", "Rol", "Estado", "MFA"]
        sheet.append(headers)
        for u in usuarios:
            sheet.append([
                u.username, u.email, u.nombre, u.apellido,
                u.get_rol_display(), u.get_estado_display(), u.get_mfa_habilitado_display()
            ])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="usuarios.xlsx"'
        workbook.save(response)
        return response

    # Eliminar usuario
    if request.method == "GET" and "delete_id" in request.GET:
        if not is_admin:
            messages.error(request, "No tienes permiso para eliminar usuarios.")
            return redirect("Formulario")
        if not user_can_delete_module(request.user, 'usuarios'):
            messages.error(request, "No tienes permiso para eliminar usuarios.")
            return redirect("Formulario")
        usuario_id = request.GET.get("delete_id")
        try:
            usuario = Usuario.objects.get(pk=usuario_id)
            username = usuario.username
            try:
                auth_user = User.objects.get(username=username)
                auth_user.delete()
            except User.DoesNotExist:
                pass
            usuario.delete()
            messages.success(request, "Usuario eliminado correctamente.")
        except Exception as e:
            messages.error(request, f"Error al eliminar usuario: {str(e)}")
        return redirect("Formulario")

    # Eliminar avatar (desde el mismo formulario principal)
    if request.method == "POST" and request.POST.get("delete_avatar"):
        edit_id = request.POST.get("edit_id")
        usuario = get_object_or_404(Usuario, pk=edit_id)
        if usuario.avatar:
            usuario.avatar.delete(save=False)
            usuario.avatar = None
            usuario.save(update_fields=['avatar'])
            usuario.sync_to_auth_user(request=request)
        messages.success(request, "Avatar eliminado correctamente.")
        return redirect("Formulario")

    # Editar o Crear usuario
    edit_mode = False
    edit_id = ""
    form = None

    if request.method == "GET" and "edit_id" in request.GET:
        edit_id = request.GET.get("edit_id")
        if not can_edit_own_profile(request.user, edit_id):
            messages.error(request, "Solo puedes editar tu propio perfil.")
            return redirect("Formulario")
        if not user_can_change_module(request.user, 'usuarios'):
            messages.error(request, "No tienes permiso para editar usuarios.")
            return redirect("Formulario")
        instance = get_object_or_404(Usuario, pk=edit_id)
        form = UsuarioForm(instance=instance)
        edit_mode = True

    elif request.method == "POST" and not request.POST.get("delete_avatar"):
        edit_id = request.POST.get("edit_id")
        if edit_id:
            if not can_edit_own_profile(request.user, edit_id):
                messages.error(request, "Solo puedes editar tu propio perfil.")
                return redirect("Formulario")
            if not user_can_change_module(request.user, 'usuarios'):
                messages.error(request, "No tienes permiso para editar usuarios.")
                return redirect("Formulario")
        else:
            if not is_admin or not user_can_add_module(request.user, 'usuarios'):
                messages.error(request, "No tienes permiso para crear usuarios.")
                return redirect("Formulario")

        instance = get_object_or_404(Usuario, pk=edit_id) if edit_id else None

        if is_admin:
            form = UsuarioForm(request.POST, request.FILES, instance=instance)
        else:
            POST_data = request.POST.copy()
            if instance:
                POST_data['avatar'] = instance.avatar
            form = UsuarioForm(POST_data, instance=instance)

        edit_mode = bool(edit_id)

        if not is_admin and instance:
            original_rol = instance.rol
            if form.is_valid():
                usuario = form.save(commit=False)
                usuario.rol = original_rol
                password = form.cleaned_data.get('password')
                if password:
                    usuario.set_password(password)
                usuario.save()
                usuario.sync_to_auth_user(request=request)
                messages.success(request, "Usuario actualizado correctamente.")
                return redirect("Formulario")
        else:
            if form.is_valid():
                usuario = form.save(commit=False)
                password = form.cleaned_data.get('password')
                if password:
                    usuario.set_password(password)
                usuario.save()
                usuario.sync_to_auth_user(request=request)
                messages.success(request, "Usuario guardado correctamente.")
                return redirect("Formulario")
    else:
        if is_admin and user_can_add_module(request.user, 'usuarios'):
            form = UsuarioForm()
        else:
            form = None
        edit_mode = False
        edit_id = ""

    show_form = is_admin or edit_mode

    return render(request, "dispositivos/formularioUsuario.html", {
        "visitas": visitas,
        "form": form,
        "usuarios": usuarios_page,
        "edit_mode": edit_mode,
        "edit_id": edit_id,
        "is_admin": is_admin,
        "is_analyst": is_analyst,
        "can_assign_roles": can_assign_roles(request.user),
        "show_form": show_form,
        "can_add": is_admin and user_can_add_module(request.user, 'usuarios'),
        "can_edit": user_can_change_module(request.user, 'usuarios'),
        "can_delete": is_admin and user_can_delete_module(request.user, 'usuarios'),
        "pag_size": pag_size,
        "order": order,
        "order_by": order_by,
    })



# -------------------------------------------------------------
# PRODUCTOS
# -------------------------------------------------------------
@login_required
@require_module_permission('productos', 'view')
def gestionProductos(request):
    visitas = request.session.get("visitas", 0)
    request.session['visitas'] = visitas + 1

    role_name = get_user_role_name(request.user)

    # --- ELIMINAR ---
    delete_id = request.GET.get('delete_id')
    if delete_id and request.method == 'GET':
        if not user_can_delete_module(request.user, 'productos'):
            messages.error(request, "No tienes permiso para eliminar productos.")
            return redirect('Productos')
        producto = get_object_or_404(Producto, pk=delete_id)
        producto.delete()
        messages.success(request, "Producto eliminado correctamente.")
        return redirect('Productos')

    # --- EDITAR (Cargar producto a form) ---
    edit_id = request.GET.get('edit_id')
    if edit_id:
        if not user_can_change_module(request.user, 'productos'):
            messages.error(request, "No tienes permiso para editar productos.")
            return redirect('Productos')
        producto_obj = get_object_or_404(Producto, pk=edit_id)
        form = ProductoForm(instance=producto_obj)
        edit_mode = True
    else:
        producto_obj = None
        form = ProductoForm()
        edit_mode = False

    # --- CREAR/ACTUALIZAR (POST) ---
    if request.method == "POST":
        if 'edit_id' in request.POST:
            if not user_can_change_module(request.user, 'productos'):
                messages.error(request, "No tienes permiso para editar productos.")
                return redirect('Productos')
            instance = get_object_or_404(Producto, pk=request.POST['edit_id'])
            form = ProductoForm(request.POST, instance=instance)
            edit_mode = True
        else:
            if not user_can_add_module(request.user, 'productos'):
                messages.error(request, "No tienes permiso para crear productos.")
                return redirect('Productos')
            form = ProductoForm(request.POST)
            edit_mode = False
        if form.is_valid():
            form.save()
            messages.success(request, "Producto guardado correctamente.")
            return redirect('Productos')

    search = request.GET.get('buscar', '')
    productos = Producto.objects.all()

    # Filtrado por búsqueda
    if search:
        productos = productos.filter(
            Q(nombre__icontains=search) | Q(sku__icontains=search) | Q(categoria__icontains=search)
        )

    # Exportar a Excel
    if "export_excel" in request.GET:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Productos"
        headers_excel = [
            "SKU", "Nombre", "Categoría", "UOM Compra", "UOM Venta", "Conversión",
            "IVA", "Stock mínimo", "Stock máximo", "Punto reorden",
            "Perecible?", "Control por lote", "Control por serie",
            "Lote", "Marca", "Modelo", "Descripción",
            "Imagen URL", "Ficha técnica URL", "Stock actual",
            "Alerta bajo stock", "Alerta por vencer"
        ]
        sheet.append(headers_excel)
        for p in productos:
            sheet.append([
                p.sku,
                p.nombre,
                p.categoria,
                p.uom_compra,
                p.uom_venta,
                p.factor_conversion,
                p.impuesto_iva,
                p.stock_minimo,
                getattr(p, "stock_maximo", ""),
                getattr(p, "punto_reorden", ""),
                "Sí" if p.perishable else "No",
                "Sí" if getattr(p, "control_por_lote", False) else "No",
                "Sí" if getattr(p, "control_por_serie", False) else "No",
                p.lote,
                getattr(p, "marca", ""),
                getattr(p, "modelo", ""),
                getattr(p, "descripcion", ""),
                getattr(p, "imagen_url", ""),
                getattr(p, "ficha_tecnica_url", ""),
                getattr(p, "stock_actual", ""),
                getattr(p, "alerta_bajo_stock", ""),
                getattr(p, "alerta_por_vencer", "")
            ])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="productos.xlsx"'
        workbook.save(response)
        return response

    # ========== ORDENAMIENTO ==========
    VALID_FIELDS = [
        'sku', 'nombre', 'categoria', 'uom_compra', 'uom_venta',
        'factor_conversion', 'impuesto_iva', 'stock_minimo', 'perishable',
        'lote'
    ]
    order = request.GET.get('order', request.session.get('productos_order', 'asc'))
    order_by = request.GET.get('order_by', request.session.get('productos_order_by', 'nombre'))
    if order_by not in VALID_FIELDS:
        order_by = 'nombre'
    request.session['productos_order'] = order
    request.session['productos_order_by'] = order_by

    if order == 'desc':
        productos = productos.order_by(f'-{order_by}')
    else:
        productos = productos.order_by(order_by)

    # Paginador
    pag_size = request.GET.get('pag_size') or request.session.get('productos_pag_size', '15')
    if pag_size not in ['5', '15', '30', '100']:
        pag_size = '15'
    request.session['productos_pag_size'] = pag_size

    from django.core.paginator import Paginator
    paginator = Paginator(productos, int(pag_size))
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    can_add = user_can_add_module(request.user, 'productos')
    can_change = user_can_change_module(request.user, 'productos')
    can_delete = user_can_delete_module(request.user, 'productos')

    return render(request, "dispositivos/gestionProductos.html", {
        "visitas": visitas,
        "productos": page_obj,
        "buscar": search,
        "can_add": can_add,
        "can_edit": can_change,
        "can_delete": can_delete,
        "role_name": role_name,
        "order": order,
        "order_by": order_by,
        "pag_size": pag_size,
        "form": form,
        "edit_mode": edit_mode,
        "edit_id": edit_id,
        "categorias": CATEGORIAS,
    })

# -------------------------------------------------------------
# PROVEEDORES
# -------------------------------------------------------------
@login_required
@require_module_permission('proveedores', 'view')
def gestionProveedores(request):
    visitas = request.session.get("visitas", 0)
    request.session['visitas'] = visitas + 1

    role_name = get_user_role_name(request.user)
    buscar = request.GET.get('buscar', '')
    pais = request.GET.get('pais', '')

    proveedores = Proveedor.objects.all()
    productos = Producto.objects.all().order_by("nombre")

    if buscar:
        proveedores = proveedores.filter(
            Q(rut_nif__icontains=buscar) | Q(razon_social__icontains=buscar)
        )
    if pais:
        proveedores = proveedores.filter(pais__iexact=pais)

    # ========== ORDENAMIENTO ==========
    VALID_FIELDS = ['rut_nif', 'razon_social', 'estado', 'email', 'pais']
    order = request.GET.get('order', request.session.get('proveedores_order', 'asc'))
    order_by = request.GET.get('order_by', request.session.get('proveedores_order_by', 'razon_social'))
    
    if order_by not in VALID_FIELDS:
        order_by = 'razon_social'
    
    request.session['proveedores_order'] = order
    request.session['proveedores_order_by'] = order_by
    
    if order == 'desc':
        proveedores = proveedores.order_by(f'-{order_by}')
    else:
        proveedores = proveedores.order_by(order_by)

    # Paginador
    pag_size = request.GET.get('pag_size') or request.session.get('proveedores_pag_size', '15')
    if pag_size not in ['5', '15', '30', '100']:
        pag_size = '15'
    request.session['proveedores_pag_size'] = pag_size
    
    paginator = Paginator(proveedores, int(pag_size))
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Exportar a Excel
    if "export_excel" in request.GET:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Proveedores"
        headers_excel = ["RUT/NIF", "Razón Social", "Estado", "Email", "País"]
        sheet.append(headers_excel)
        for p in proveedores:
            sheet.append([p.rut_nif, p.razon_social, p.get_estado_display(), p.email, p.pais])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="proveedores.xlsx"'
        workbook.save(response)
        return response

    can_add = user_can_add_module(request.user, 'proveedores')
    can_change = user_can_change_module(request.user, 'proveedores')
    can_delete = user_can_delete_module(request.user, 'proveedores')

    # --- Modo edición / creación / eliminación ---
    if request.method == "GET" and "edit_id" in request.GET:
        if not can_change:
            messages.error(request, "No tienes permiso para editar proveedores.")
            return redirect("Proveedores")
        form = ProveedorForm(instance=get_object_or_404(Proveedor, pk=request.GET.get("edit_id")))
        edit_mode = True

    elif request.method == "POST":
        edit_id = request.POST.get("edit_id")

        if edit_id and not can_change:
            messages.error(request, "No tienes permiso para editar proveedores.")
            return redirect("Proveedores")
        if not edit_id and not can_add:
            messages.error(request, "No tienes permiso para crear proveedores.")
            return redirect("Proveedores")

        instance = get_object_or_404(Proveedor, pk=edit_id) if edit_id else None
        form = ProveedorForm(request.POST, instance=instance)
        edit_mode = bool(edit_id)

        if form.is_valid():
            proveedor = form.save(commit=False)

            usuario_actual = Usuario.objects.filter(username=request.user.username).first()
            if not usuario_actual:
                usuario_actual = Usuario.objects.create(
                    username=request.user.username,
                    email=request.user.email or "",
                    nombre=request.user.first_name or request.user.username,
                    apellido=request.user.last_name or "",
                    estado="activo",
                    rol="operador_compras"
                )

            proveedor.usuario = usuario_actual
            proveedor.save()
            messages.success(request, "Proveedor guardado correctamente.")
            return redirect("Proveedores")

        else:
            messages.error(request, "Por favor, revisa los campos obligatorios antes de guardar.")

    elif request.method == "GET" and "delete_id" in request.GET:
        if not can_delete:
            messages.error(request, "No tienes permiso para eliminar proveedores.")
            return redirect("Proveedores")
        Proveedor.objects.filter(pk=request.GET.get("delete_id")).delete()
        messages.success(request, "Proveedor eliminado correctamente.")
        return redirect("Proveedores")

    else:
        form = ProveedorForm() if can_add else None
        edit_mode = False

    return render(request, "dispositivos/gestionProveedores.html", {
        "visitas": visitas,
        "form": form,
        "proveedores": page_obj,
        "productos": productos,
        "edit_mode": edit_mode,
        "edit_id": request.GET.get("edit_id", ""),
        "buscar": buscar,
        "pais": pais,
        "can_add": can_add,
        "can_edit": can_change,
        "can_delete": can_delete,
        "role_name": role_name,
        "order": order,
        "order_by": order_by,
        "pag_size": pag_size,
    })

# -------------------------------------------------------------
# PRODUCTO - PROVEEDOR (MOVIMIENTOS DE INVENTARIO)
# -------------------------------------------------------------
@login_required
@require_module_permission('producto_proveedor', 'view')
def moduloTransaccional(request):
    visitas = request.session.get("visitas", 0)
    request.session['visitas'] = visitas + 1

    role_name = get_user_role_name(request.user)
    search = request.GET.get("buscar", "")
    tipo = request.GET.get("tipo", "")
    movimientos = ProductoProveedor.objects.select_related('producto', 'proveedor').all()

    if search:
        movimientos = movimientos.filter(Q(producto__sku__icontains=search) | Q(producto__nombre__icontains=search))
    if tipo:
        movimientos = movimientos.filter(tipo_movimiento__iexact=tipo)

    # ========== ORDENAMIENTO ==========
    VALID_FIELDS = ['fecha_movimiento', 'tipo_movimiento', 'producto__sku', 
                    'producto__nombre', 'proveedor__razon_social', 'cantidad']
    order = request.GET.get('order', request.session.get('inventario_order', 'desc'))
    order_by = request.GET.get('order_by', request.session.get('inventario_order_by', 'fecha_movimiento'))
    
    if order_by not in VALID_FIELDS:
        order_by = 'fecha_movimiento'
    
    request.session['inventario_order'] = order
    request.session['inventario_order_by'] = order_by
    
    if order == 'desc':
        movimientos = movimientos.order_by(f'-{order_by}')
    else:
        movimientos = movimientos.order_by(order_by)

    # --- Paginador robusto ---
    pag_size = request.GET.get('pag_size') or request.session.get('inventario_pag_size', '15')
    if pag_size not in ['5', '15', '30', '100']:
        pag_size = '15'
    request.session['inventario_pag_size'] = pag_size

    paginator = Paginator(movimientos, int(pag_size))
    page_number = request.GET.get('page')
    movimientos_page = paginator.get_page(page_number)

    # Exportar a Excel (siempre exporta TODO el queryset filtrado, no solo la página)
    if "export_excel" in request.GET:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Movimientos"
        headers = ["Fecha", "Tipo", "Producto", "Proveedor", "Cantidad"]
        sheet.append(headers)
        for m in movimientos:
            sheet.append([
                timezone.localtime(m.fecha_movimiento).strftime("%Y-%m-%d %H:%M"),
                m.get_tipo_movimiento_display(),
                m.producto.nombre,
                m.proveedor.razon_social,
                m.cantidad,
            ])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="movimientos.xlsx"'
        workbook.save(response)
        return response

    # Verificar permisos CRUD
    can_add = user_can_add_module(request.user, 'producto_proveedor')
    can_change = user_can_change_module(request.user, 'producto_proveedor')
    can_delete = user_can_delete_module(request.user, 'producto_proveedor')

    # EDITAR
    if request.method == "GET" and "edit_id" in request.GET:
        if not can_change:
            messages.error(request, "No tienes permiso para editar movimientos.")
            return redirect("Transaccional")
        form = ProductoProveedorForm(instance=get_object_or_404(ProductoProveedor, pk=request.GET.get("edit_id")))
        edit_mode = True

    # GUARDAR / ACTUALIZAR
    elif request.method == "POST":
        edit_id = request.POST.get("edit_id")
        if edit_id and not can_change:
            messages.error(request, "No tienes permiso para editar movimientos.")
            return redirect("Transaccional")
        if not edit_id and not can_add:
            messages.error(request, "No tienes permiso para crear movimientos.")
            return redirect("Transaccional")

        instance = get_object_or_404(ProductoProveedor, pk=edit_id) if edit_id else None
        form = ProductoProveedorForm(request.POST, instance=instance)
        edit_mode = bool(edit_id)

        if form.is_valid():
            movimiento = form.save(commit=False)

            # Control de campos opcionales (evitar null vacíos)
            if not movimiento.lote:
                movimiento.lote = None
            if not movimiento.serie:
                movimiento.serie = None
            if not movimiento.doc_referencia:
                movimiento.doc_referencia = None
            if not movimiento.motivo:
                movimiento.motivo = None
            if not movimiento.observaciones:
                movimiento.observaciones = None

            movimiento.save()
            messages.success(request, "Movimiento guardado correctamente.")
            return redirect("Transaccional")
        else:
            messages.error(request, "Error al guardar el movimiento. Verifica los campos.")

    # ELIMINAR
    elif request.method == "GET" and "delete_id" in request.GET:
        if not can_delete:
            messages.error(request, "No tienes permiso para eliminar movimientos.")
            return redirect("Transaccional")
        ProductoProveedor.objects.filter(pk=request.GET.get("delete_id")).delete()
        messages.success(request, "Movimiento eliminado correctamente.")
        return redirect("Transaccional")

    else:
        form = ProductoProveedorForm() if can_add else None
        edit_mode = False

    movimientos_hoy = ProductoProveedor.objects.filter(
        fecha_movimiento__date=timezone.now().date()
    ).count()

    return render(request, "dispositivos/moduloTransaccional.html", {
        "visitas": visitas,
        "form": form,
        "movimientos": movimientos_page,
        "pag_size": pag_size,
        "movimientos_hoy": movimientos_hoy,
        "productos_count": Producto.objects.count(),
        "edit_mode": edit_mode,
        "edit_id": request.GET.get("edit_id", ""),
        "can_add": can_add,
        "can_edit": can_change,
        "can_delete": can_delete,
        "role_name": role_name,
        "order": order,
        "order_by": order_by,
    })
    
# PERFIL :3
@login_required
def perfilusuario(request):
    user = request.user
    usuario_qs = Usuario.objects.filter(username=user.username)
    if not usuario_qs.exists() and user.email:
        usuario_qs = Usuario.objects.filter(email=user.email)
    usuario = usuario_qs.first()
    if not usuario:
        usuario = Usuario.objects.create(
            username=user.username,
            email=user.email,
            nombre=user.first_name if user.first_name else user.username,
            apellido=user.last_name or "",
            rol='operador_ventas',
            estado='activo',
        )

    perfil_form = PerfilUsuarioForm(instance=usuario)
    pass_form = PasswordChangeForm(usuario=usuario)

    if request.method == "POST":
        form_type = request.POST.get("form_type")
        # --- Avatar solo (y borrar) ---
        if form_type == "avatar":
            if "delete_avatar" in request.POST:
                usuario.avatar.delete(save=False)
                usuario.avatar = None
                usuario.save(update_fields=['avatar'])
                usuario.sync_to_auth_user(request=request)
                messages.success(request, "Foto de perfil eliminada correctamente.")
                return redirect("perfil_usuario")
            perfil_form = PerfilUsuarioForm(request.POST, request.FILES, instance=usuario)
            if perfil_form.is_valid():
                if perfil_form.cleaned_data.get("avatar"):
                    usuario.avatar = perfil_form.cleaned_data["avatar"]
                    usuario.save(update_fields=['avatar'])
                    usuario.sync_to_auth_user(request=request)
                    messages.success(request, "Foto de perfil actualizada correctamente.")
                else:
                    messages.error(request, "Debes seleccionar una imagen válida.")
            return redirect("perfil_usuario")
        # --- Datos personales solo ---
        elif form_type == "datos":
            perfil_form = PerfilUsuarioForm(request.POST, instance=usuario)
            if perfil_form.is_valid():
                usuario.nombre = perfil_form.cleaned_data["nombre"]
                usuario.apellido = perfil_form.cleaned_data["apellido"]
                usuario.email = perfil_form.cleaned_data["email"]
                usuario.save(update_fields=['nombre', 'apellido', 'email'])
                usuario.sync_to_auth_user(request=request)
                messages.success(request, "Perfil actualizado correctamente.")
            else:
                messages.error(request, "Corrige los errores del formulario.")
            return redirect("perfil_usuario")
        # --- Cambio de contraseña robusto ---
        elif form_type == "password":
            pass_form = PasswordChangeForm(request.POST, usuario=usuario)
            if pass_form.is_valid():
                new_pass = pass_form.cleaned_data["new_password1"]
                usuario.set_password(new_pass)
                usuario.save(update_fields=['password'])
                user.set_password(new_pass)
                user.save()
                update_session_auth_hash(request, user)
                usuario.sync_to_auth_user(request=request)
                messages.success(request, "Contraseña actualizada correctamente.")
            else:
                # Muestra todos los errores del formulario
                error_msgs = []
                for field, errors in pass_form.errors.items():
                    for error in errors:
                        error_msgs.append(f"{pass_form.fields[field].label if field in pass_form.fields else field}: {error}")
                messages.error(request, "Hubo un error al cambiar la contraseña.<br>" + "<br>".join(error_msgs))
            return redirect("perfil_usuario")

    return render(request, "dispositivos/perfilusuario.html", {
        "perfil_form": PerfilUsuarioForm(instance=usuario),
        "pass_form": PasswordChangeForm(usuario=usuario),
        "usuario": usuario,
        "user": user,
    })

